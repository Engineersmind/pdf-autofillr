"""XLSX reader using openpyxl."""
from __future__ import annotations
from uploaddocument.readers.base_reader import DocumentReader


class XlsxReader(DocumentReader):
    def read(self, file_path: str) -> str:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            rows = [f"--- Sheet: {name} ---"]
            for row in wb[name].iter_rows(values_only=True):
                row_text = " | ".join(str(c) if c is not None else "" for c in row).strip()
                if row_text:
                    rows.append(row_text)
            if len(rows) > 1:
                sheets.append("\n".join(rows))
        return "\n\n".join(sheets)
